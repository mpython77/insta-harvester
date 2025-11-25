'''"""
Instagram Scraper - Excel Export Utilities
Real-time Excel export with pandas and openpyxl
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging


class ExcelExporter:
    """
    Real-time Excel exporter

    Columns:
    1. Post URL
    2. Type (Post/Reel)
    3. Tagged Accounts
    4. Likes Count
    5. Post Date
    6. Scraping Date/Time
    """

    def __init__(self, filename: str, logger: Optional[logging.Logger] = None):
        """
        Initialize Excel exporter

        Args:
            filename: Output Excel filename
            logger: Logger instance
        """
        self.filename = Path(filename)
        self.logger = logger or logging.getLogger(__name__)

        # Use list for O(1) append performance, convert to DataFrame when writing
        self.rows: List[Dict[str, Any]] = []

        # Column names
        self.columns = [
            'Post URL',
            'Type',
            'Tagged Accounts',
            'Likes Count',
            'Post Date',
            'Scraping Date/Time'
        ]

        # Create empty file
        self._create_file()

        self.logger.info(f"Excel exporter initialized: {self.filename}")

    def _create_file(self) -> None:
        """Create initial Excel file with headers"""
        try:
            # Create empty DataFrame with just headers
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.filename, index=False, engine='openpyxl')
            self.logger.debug(f"Created Excel file: {self.filename}")
        except Exception as e:
            self.logger.error(f"Failed to create Excel file: {e}")
            raise

    def add_row(
        self,
        post_url: str,
        tagged_accounts: List[str],
        likes: str,
        post_date: str,
        content_type: str = 'Post'
    ) -> None:
        """
        Add a single row to Excel (real-time)

        Args:
            post_url: Post URL
            tagged_accounts: List of tagged usernames
            likes: Likes count
            post_date: Post timestamp
            content_type: Content type ('Post' or 'Reel')
        """
        try:
            # Format tags
            tags_str = ', '.join(tagged_accounts) if tagged_accounts else 'No tags'

            # Current datetime
            scraping_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Append to list (O(1) operation - fast!)
            row_dict = {
                'Post URL': post_url,
                'Type': content_type,
                'Tagged Accounts': tags_str,
                'Likes Count': likes,
                'Post Date': post_date,
                'Scraping Date/Time': scraping_time
            }
            self.rows.append(row_dict)

            # Convert to DataFrame and write to file (real-time)
            df = pd.DataFrame(self.rows, columns=self.columns)
            df.to_excel(self.filename, index=False, engine='openpyxl')

            self.logger.debug(f"Added row to Excel [{content_type}]: {post_url}")

        except Exception as e:
            self.logger.error(f"Failed to add row to Excel: {e}")

    def add_multiple_rows(self, data: List[Dict[str, Any]]) -> None:
        """
        Add multiple rows at once

        Args:
            data: List of dictionaries with post data
        """
        for item in data:
            self.add_row(
                post_url=item.get('url', 'N/A'),
                tagged_accounts=item.get('tagged_accounts', []),
                likes=item.get('likes', 'N/A'),
                post_date=item.get('timestamp', 'N/A'),
                content_type=item.get('content_type', 'Post')
            )

    def get_row_count(self) -> int:
        """Get current number of rows"""
        return len(self.rows)

    def finalize(self) -> None:
        """Finalize Excel file (optional cleanup)"""
        try:
            # Auto-adjust column widths
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(self.filename)
            ws = wb.active

            # Auto-width for all columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Max 50
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.filename)
            self.logger.info(f"Excel file finalized: {self.filename}")

        except Exception as e:
            self.logger.warning(f"Failed to auto-adjust columns: {e}")
            '''

"""
Instagram Scraper - Excel Export Utilities
Real-time Excel export with pandas and openpyxl
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging


class ExcelExporter:
    """
    Real-time Excel exporter

    Columns:
    1. Post URL
    2. Type (Post/Reel)
    3. Tagged Account (har bir tag alohida qatorda)
    4. Likes Count
    5. Post Date
    6. Scraping Date/Time
    """

    def __init__(
        self, 
        filename: str, 
        logger: Optional[logging.Logger] = None, 
        batch_size: int = 10,
        separate_tags: bool = True
    ):
        """
        Initialize Excel exporter

        Args:
            filename: Output Excel filename
            logger: Logger instance
            batch_size: Har nechta rowda saqlash (default: 10)
            separate_tags: Agar True bo'lsa, har bir tag alohida qatorda (default: True)
        """
        self.filename = Path(filename)
        self.logger = logger or logging.getLogger(__name__)
        self.batch_size = batch_size
        self.separate_tags = separate_tags

        self.rows: List[Dict[str, Any]] = []

        column_name = 'Tagged Account' if separate_tags else 'Tagged Accounts'
        self.columns = [
            'Post URL',
            'Type',
            column_name,
            'Likes Count',
            'Post Date',
            'Scraping Date/Time'
        ]

        self._create_file()

        self.logger.info(f"Excel exporter initialized: {self.filename}")

    def _create_file(self) -> None:
        """Create initial Excel file with headers"""
        try:
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.filename, index=False, engine='openpyxl')
            self.logger.debug(f"Created Excel file: {self.filename}")
        except Exception as e:
            self.logger.error(f"Failed to create Excel file: {e}")
            raise

    def _write_to_excel(self) -> None:
        """Write current rows to Excel file"""
        try:
            df = pd.DataFrame(self.rows, columns=self.columns)
            df.to_excel(self.filename, index=False, engine='openpyxl')
            self.logger.debug(f"Saved {len(self.rows)} rows to Excel")
        except Exception as e:
            self.logger.error(f"Failed to write to Excel: {e}")

    def add_row(
        self,
        post_url: str,
        tagged_accounts: List[str],
        likes: str,
        post_date: str,
        content_type: str = 'Post'
    ) -> None:
        """
        Add a single row to Excel (real-time)

        Args:
            post_url: Post URL
            tagged_accounts: List of tagged usernames
            likes: Likes count
            post_date: Post timestamp
            content_type: Content type ('Post' or 'Reel')
        """
        try:
            scraping_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            if self.separate_tags:
                # HAR BIR TAG ALOHIDA QATORDA
                if not tagged_accounts:
                    row_dict = {
                        'Post URL': post_url,
                        'Type': content_type,
                        'Tagged Account': 'No tags',
                        'Likes Count': likes,
                        'Post Date': post_date,
                        'Scraping Date/Time': scraping_time
                    }
                    self.rows.append(row_dict)
                else:
                    for tag in tagged_accounts:
                        row_dict = {
                            'Post URL': post_url,
                            'Type': content_type,
                            'Tagged Account': tag,
                            'Likes Count': likes,
                            'Post Date': post_date,
                            'Scraping Date/Time': scraping_time
                        }
                        self.rows.append(row_dict)
                
                self.logger.debug(f"Added {len(tagged_accounts) if tagged_accounts else 1} rows [{content_type}]: {post_url}")
            
            else:
                # ESKI LOGIKA: Barcha taglar bitta qatorda
                tags_str = ', '.join(tagged_accounts) if tagged_accounts else 'No tags'
                
                row_dict = {
                    'Post URL': post_url,
                    'Type': content_type,
                    'Tagged Accounts': tags_str,
                    'Likes Count': likes,
                    'Post Date': post_date,
                    'Scraping Date/Time': scraping_time
                }
                self.rows.append(row_dict)
                
                self.logger.debug(f"Added row to Excel [{content_type}]: {post_url}")

            # Har batch_size ta rowda saqlash
            if len(self.rows) % self.batch_size == 0:
                self._write_to_excel()

        except Exception as e:
            self.logger.error(f"Failed to add row to Excel: {e}")

    def add_multiple_rows(self, data: List[Dict[str, Any]]) -> None:
        """
        Add multiple rows at once

        Args:
            data: List of dictionaries with post data
        """
        for item in data:
            self.add_row(
                post_url=item.get('url', 'N/A'),
                tagged_accounts=item.get('tagged_accounts', []),
                likes=item.get('likes', 'N/A'),
                post_date=item.get('timestamp', 'N/A'),
                content_type=item.get('content_type', 'Post')
            )

    def get_row_count(self) -> int:
        """Get current number of rows"""
        return len(self.rows)

    def finalize(self) -> None:
        """Finalize Excel file (optional cleanup)"""
        try:
            # Qolgan ma'lumotlarni saqlash
            if len(self.rows) % self.batch_size != 0:
                self._write_to_excel()

            # Auto-adjust column widths
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(self.filename)
            ws = wb.active

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.filename)
            self.logger.info(f"Excel file finalized: {self.filename}")

        except Exception as e:
            self.logger.warning(f"Failed to auto-adjust columns: {e}")

