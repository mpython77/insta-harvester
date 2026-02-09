"""
Instagram Scraper - Unified Export Utilities
Professional real-time export to Excel and JSON with strict typing.

This module consolidates all export functionality:
- Profile/Post Data Export (Excel)
- Comment Data Export (JSON/Excel)
- Real-time streaming support
"""

import json
import csv
import os
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Union, Protocol, TypedDict
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .config import ScraperConfig
from .models import Comment, CommentData
# Using forward reference for PostCommentsData to avoid circular imports if possible, 
# or strictly importing if Models are clean. 
# Assuming models.py might need update or we use Dict for flexibility in exporter.

# TypedDicts for better type safety
class PostDataDict(TypedDict, total=False):
    url: str
    tagged_accounts: List[str]
    likes: str
    timestamp: str
    content_type: str

class CommentDataDict(TypedDict, total=False):
    id: str
    username: str
    text: str
    date: str
    likes: int
    replies: int
    url: str
    is_reply: bool
    parent_id: Optional[str]

# ==================== Interfaces ====================

class BaseExporter(Protocol):
    """Protocol for all exporters"""
    def add_row(self, data: Dict[str, Any]) -> None: ...
    def finalize(self) -> None: ...

# ==================== Excel Exporters ====================

class ExcelExporter:
    """
    Real-time Excel exporter for Posts/Reels data.
    
    Features:
    - Real-time writing (batch based)
    - Auto-adjust column widths
    - Tag separation support
    """
    
    def __init__(
        self, 
        filename: str, 
        logger: Optional[logging.Logger] = None, 
        config: Optional[ScraperConfig] = None
    ):
        self.filename = Path(filename)
        self.logger = logger or logging.getLogger(__name__)
        self.config = config or ScraperConfig()
        
        self.batch_size = 10
        self.separate_tags = True
        self.rows: List[Dict[str, Any]] = []
        
        column_name = 'Tagged Account' if self.separate_tags else 'Tagged Accounts'
        self.columns = [
            'Post URL', 'Type', column_name, 
            'Likes Count', 'Post Date', 'Scraping Date/Time'
        ]
        
        self._create_file()
        self.logger.info(f"Excel exporter initialized: {self.filename}")

    def _create_file(self) -> None:
        """Create initial Excel file with headers"""
        try:
            # Check if file exists to avoid overwriting if resuming (optional logic could go here)
            # For now, we overwrite as per original logic, or start fresh.
            df = pd.DataFrame(columns=self.columns)
            df.to_excel(self.filename, index=False, engine='openpyxl')
        except Exception as e:
            self.logger.error(f"Failed to create Excel file: {e}")
            raise

    def _write_to_excel(self) -> None:
        """Write current rows to Excel file"""
        try:
            df = pd.DataFrame(self.rows, columns=self.columns)
            # engine='openpyxl' allows working with existing files usually, 
            # but pandas to_excel overwrites by default. 
            # To append efficiently without re-reading, we can use openpyxl directly or append mode.
            # However, original implementation used pandas overwrite for simplicity on batch.
            # Let's optimize to use openpyxl append for true streaming/performance if file is large.
            
            # Using append mode with openpyxl for performance
            if not self.rows:
                return

            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            for row in self.rows:
                # Map dict to list based on columns order
                row_values = []
                for col in self.columns:
                    row_values.append(row.get(col, ''))
                ws.append(row_values)
                
            wb.save(self.filename)
            self.rows = [] # Clear buffer
            self.logger.debug("Saved batch to Excel")
            
        except Exception as e:
            self.logger.error(f"Failed to write to Excel: {e}")

    def add_row(
        self,
        post_url: str,
        tagged_accounts: List[str],
        likes: str,
        post_date: str,
        content_type: str = 'Post'
    ) -> None:
        """Add a single row (or multiple if tags separated)"""
        try:
            scraping_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            if self.separate_tags:
                if not tagged_accounts:
                    self._buffer_row(post_url, content_type, 'No tags', likes, post_date, scraping_time)
                else:
                    for tag in tagged_accounts:
                        self._buffer_row(post_url, content_type, tag, likes, post_date, scraping_time)
            else:
                tags_str = ', '.join(tagged_accounts) if tagged_accounts else 'No tags'
                self._buffer_row(post_url, content_type, tags_str, likes, post_date, scraping_time)
                
        except Exception as e:
            self.logger.error(f"Failed to add row: {e}")

    def _buffer_row(self, url, c_type, tags, likes, date, time_str):
        column_name = 'Tagged Account' if self.separate_tags else 'Tagged Accounts'
        self.rows.append({
            'Post URL': url,
            'Type': c_type,
            column_name: tags,
            'Likes Count': likes,
            'Post Date': date,
            'Scraping Date/Time': time_str
        })
        
        if len(self.rows) >= self.batch_size:
            self._write_to_excel()

    def add_multiple_rows(self, data: List[Dict[str, Any]]) -> None:
        for item in data:
            self.add_row(
                post_url=item.get('url', 'N/A'),
                tagged_accounts=item.get('tagged_accounts', []),
                likes=item.get('likes', 'N/A'),
                post_date=item.get('timestamp', 'N/A'),
                content_type=item.get('content_type', 'Post')
            )

    def finalize(self) -> None:
        """Save remaining rows and adjust column widths"""
        if self.rows:
            self._write_to_excel()
        self._auto_adjust_columns()
        self.logger.info(f"Excel finalized: {self.filename}")

    def _auto_adjust_columns(self):
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception: pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            wb.save(self.filename)
        except Exception as e:
            self.logger.warning(f"Column adjustment failed: {e}")

# ==================== Comments Exporters ====================

class CommentsExporter:
    """
    Specialized exporter for Comments with support for nested replies.
    Can export to JSON (hierarchical) and Excel (flat).
    """
    
    def __init__(
        self,
        username: str,
        logger: Optional[logging.Logger] = None,
        config: Optional[ScraperConfig] = None,
        export_json: bool = True,
        export_excel: bool = True
    ):
        self.username = username
        self.logger = logger or logging.getLogger(__name__)
        self.config = config or ScraperConfig()
        self.export_json = export_json
        self.export_excel = export_excel
        
        self.comments_data: List[Any] = [] # Stores PostCommentsData-like objects
        self.excel_rows: List[Dict[str, Any]] = []

    def add_post_comments(self, post_comments: Any) -> None:
        """Add all comments from a post data object"""
        self.comments_data.append(post_comments)
        
        if self.export_excel:
            # Flatten for Excel
            base_info = {
                'Post URL': post_comments.post_url,
                'Post ID': post_comments.post_id
            }
            
            # Use generator if available or iterate list
            iterator = getattr(post_comments, 'get_all_comments_flat', None)
            if iterator:
                for comment in iterator():
                    self._add_to_excel_buffer(base_info, comment)
            else:
                # Fallback if get_all_comments_flat missing
                for comment in post_comments.comments:
                    self._add_to_excel_buffer(base_info, comment, is_reply=False)
                    for reply in comment.replies:
                         self._add_to_excel_buffer(base_info, reply, is_reply=True, parent_id=comment.id)

    def _add_to_excel_buffer(self, base_info: Dict, comment: Any, is_reply: bool = False, parent_id: str = None):
        # Handle both flat definition (from generator) and recursive manual
        # If comment object has 'is_reply' attribute, trust it
        is_reply_val = getattr(comment, 'is_reply', is_reply)
        parent_id_val = getattr(comment, 'parent_id', parent_id)
        
        row = base_info.copy()
        row.update({
            'Comment ID': comment.id,
            'Author Username': comment.author.username,
            'Author Verified': comment.author.is_verified,
            'Comment Text': comment.text,
            'Likes Count': comment.likes_count,
            'Reply Count': comment.reply_count,
            'Timestamp': comment.timestamp,
            'Timestamp ISO': comment.timestamp_iso,
            'Is Reply': is_reply_val,
            'Parent Comment ID': parent_id_val,
            'Comment URL': comment.permalink,
            'Scraped At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        self.excel_rows.append(row)

    def finalize(self):
        if self.export_json:
            self._save_json()
        if self.export_excel:
            self._save_excel()

    def _save_json(self):
        filename = self.config.comments_json_filename_pattern.format(
            username=self.username, 
            post_id="BATCH" # Generalized for batch
        )
        # Verify if pattern has post_id, if we are saving multiple posts, maybe separate files? 
        # Original logic implies one file per post usually, or one big file.
        # Let's save one big file for batch export
        if "{post_id}" in self.config.comments_json_filename_pattern:
             filename = f"comments_{self.username}_batch.json"
        
        try:
            data = [p.to_dict() for p in self.comments_data]
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.logger.info(f"Saved JSON comments to {filename}")
        except Exception as e:
            self.logger.error(f"Failed to save JSON: {e}")

    def _save_excel(self):
        if not self.excel_rows:
            return
            
        filename = self.config.comments_excel_filename_pattern.format(username=self.username)
        try:
            df = pd.DataFrame(self.excel_rows)
            # Reorder columns if config present
            if hasattr(self.config, 'comments_excel_columns'):
                # Only select columns that exist in df
                cols = [c for c in self.config.comments_excel_columns if c in df.columns]
                if cols:
                    df = df[cols]
            
            df.to_excel(filename, index=False)
            self.logger.info(f"Saved Excel comments to {filename}")
        except Exception as e:
            self.logger.error(f"Failed to save Excel: {e}")

class RealTimeCommentsExporter(CommentsExporter):
    """
    Writes comments to disk immediately.
    Prevent data loss during long scrapes.
    """
    def __init__(self, *args, batch_size=5, **kwargs):
        super().__init__(*args, **kwargs)
        self.batch_size = batch_size
        self.current_rows: List[Dict] = []
        self.wb: Optional[Workbook] = None
        self.ws = None
        
        self.excel_filename = self.config.comments_excel_filename_pattern.format(username=self.username)
        if self.export_excel:
            self._init_excel_stream()

    def _init_excel_stream(self):
        if os.path.exists(self.excel_filename):
            try:
                self.wb = openpyxl.load_workbook(self.excel_filename)
                self.ws = self.wb.active
            except Exception:
                self.wb = Workbook()
                self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            # Header
            self.ws.append(self.config.comments_excel_columns)
            self.wb.save(self.excel_filename)

    def on_comment_scraped(self, count: int, comment: Any, post_url: str, post_id: str):
        """Callback for real-time saving"""
        # Prepare row
        is_reply = getattr(comment, 'is_reply', False)
        parent_id = getattr(comment, 'parent_id', None)
        
        row = {
            'Post URL': post_url,
            'Post ID': post_id,
            'Comment ID': comment.id,
            'Author Username': comment.author.username,
            'Author Verified': comment.author.is_verified,
            'Comment Text': comment.text,
            'Likes Count': comment.likes_count,
            'Reply Count': comment.reply_count,
            'Timestamp': comment.timestamp,
            'Timestamp ISO': comment.timestamp_iso,
            'Is Reply': is_reply,
            'Parent Comment ID': parent_id,
            'Comment URL': comment.permalink,
            'Scraped At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        self.current_rows.append(row)
        
        if len(self.current_rows) >= self.batch_size:
            self._flush_batch()

    def _flush_batch(self):
        if not self.export_excel or not self.current_rows:
            return
            
        try:
            # Re-load workbook to ensure we don't overwrite if multiple processes (basic check)
            # For strict single-process streaming, keeping open is faster, but safer to load/save for reliability
            # Optimization: Keep open, specific save.
            
            for row_dict in self.current_rows:
                row_values = []
                for col in self.config.comments_excel_columns:
                    row_values.append(row_dict.get(col, ''))
                self.ws.append(row_values)
                
            self.wb.save(self.excel_filename)
            self.current_rows = []
        except Exception as e:
            self.logger.error(f"Failed to flush batch: {e}")

    def finalize(self):
        self._flush_batch()
        super().finalize()

# ==================== Generic Streamers ====================

class StreamingJSONExporter:
    """Writes JSON Lines (JSONL) to file"""
    def __init__(self, filename: str):
        self.filename = filename
        Path(filename).parent.mkdir(parents=True, exist_ok=True)

    def append_item(self, item: Dict[str, Any]):
        with open(self.filename, 'a', encoding='utf-8') as f:
            f.write(json.dumps(item, ensure_ascii=False) + '\n')

class StreamingExcelExporter:
    """Generic Streaming Excel"""
    def __init__(self, filename: str, columns: List[str]):
        self.filename = filename
        self.columns = columns
        self._init_file()

    def _init_file(self):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.append(self.columns)
            wb.save(self.filename)

    def append_row(self, row: List[Any]):
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        ws.append(row)
        wb.save(self.filename)

# ==================== Utility Functions ====================

def export_comments_to_json(
    comments_data: Union[Any, List[Any]], 
    filename: str, 
    logger: Optional[logging.Logger] = None
) -> bool:
    """Standalone helper function"""
    try:
        data = []
        if isinstance(comments_data, list):
            data = [c.to_dict() for c in comments_data]
        else:
            data = [comments_data.to_dict()]
            
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        if logger: logger.error(f"JSON export failed: {e}")
        return False

def export_comments_to_excel(
    comments_data: Union[Any, List[Any]], 
    filename: str,
    logger: Optional[logging.Logger] = None
) -> bool:
    """Standalone helper function"""
    try:
        # Use CommentsExporter to flatten data but handle saving manually
        exporter = CommentsExporter("temp", logger=logger, export_json=False, export_excel=False)
        
        if isinstance(comments_data, list):
            for cd in comments_data:
                exporter.add_post_comments(cd)
        else:
            exporter.add_post_comments(comments_data)
            
        if not exporter.excel_rows:
            return False
            
        df = pd.DataFrame(exporter.excel_rows)
        df.to_excel(filename, index=False)
        return True
    except Exception as e:
         if logger: logger.error(f"Excel export failed: {e}")
         return False
